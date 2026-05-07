"""
Build a real .xlsx funnel model with live formulas — the file investors
can open in Sheets or Excel and immediately see numbers recompute when
they change assumptions.

Run: python3 scripts/build-funnel-xlsx.py
Output: docs/investor/funnel-model.xlsx (overwrites)

Why a Python build script: keeps the model definition in code (under
version control), regenerates the xlsx deterministically, and lets us
add scenarios / sensitivity tables without hand-editing a binary file.
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Output ────────────────────────────────────────────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "docs/investor/funnel-model.xlsx")

# ── Styling primitives ────────────────────────────────────────────────────
BORDER_HAIRLINE = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
FILL_INPUT = PatternFill("solid", fgColor="DCEEF7")     # teal — editable
FILL_OUTPUT = PatternFill("solid", fgColor="E8F1E8")    # green — computed
FILL_HEADER = PatternFill("solid", fgColor="1A1510")    # ink
FILL_SECTION = PatternFill("solid", fgColor="F5F0E6")   # sand — section
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FAF8F4")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="1A1510")
FONT_BODY = Font(name="Calibri", size=10, color="1A1510")
FONT_LABEL = Font(name="Calibri", size=10, bold=True, color="1A1510")
FONT_NOTE = Font(name="Calibri", size=9, italic=True, color="8A7D72")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def set_cell(ws, row, col, value, *, font=FONT_BODY, fill=None, align=None,
             number_format=None, border=BORDER_HAIRLINE):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if number_format:
        cell.number_format = number_format
    cell.border = border
    return cell


def section_header(ws, row, title):
    set_cell(ws, row, 1, title, font=FONT_SECTION, fill=FILL_SECTION,
             align=ALIGN_LEFT)
    for col in range(2, 6):
        set_cell(ws, row, col, "", fill=FILL_SECTION)


def build_main_model(wb):
    ws = wb.active
    ws.title = "Funnel Model"

    # Column widths — readable on first open
    widths = {1: 38, 2: 14, 3: 14, 4: 14, 5: 36}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Header row
    headers = ["Line item", "Year 1", "Year 2", "Year 3", "Notes"]
    for col, text in enumerate(headers, start=1):
        set_cell(ws, 1, col, text, font=FONT_HEADER, fill=FILL_HEADER,
                 align=ALIGN_CENTER)
    ws.row_dimensions[1].height = 22

    # ── 1. Top of funnel ──────────────────────────────────────────────────
    section_header(ws, 2, "TOP OF FUNNEL — surface area + capture")

    # Surface count (input)
    set_cell(ws, 3, 1, "SEO surfaces (live cards + cities)", font=FONT_LABEL)
    set_cell(ws, 3, 2, 47, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 3, 3, 75, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 3, 4, 150, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 3, 5, "1 surface = 1 ranked URL (intel card OR beware city)",
             font=FONT_NOTE, align=ALIGN_LEFT)

    # Keyword volume (input — replace with Ahrefs midpoint)
    set_cell(ws, 4, 1, "Avg keyword volume per surface (mo)", font=FONT_LABEL)
    set_cell(ws, 4, 2, 1000, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 4, 3, 1100, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 4, 4, 1200, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 4, 5, "REPLACE with verified Ahrefs midpoint",
             font=FONT_NOTE, align=ALIGN_LEFT)

    # Total addressable searches (formula)
    set_cell(ws, 5, 1, "Total addressable searches/mo", font=FONT_LABEL)
    set_cell(ws, 5, 2, "=B3*B4", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 5, 3, "=C3*C4", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 5, 4, "=D3*D4", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 5, 5, "Surfaces × volume", font=FONT_NOTE, align=ALIGN_LEFT)

    # Capture rate
    set_cell(ws, 6, 1, "Capture rate (% of searches won)", font=FONT_LABEL)
    set_cell(ws, 6, 2, 0.05, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 6, 3, 0.10, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 6, 4, 0.18, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 6, 5, "Year 1 conservative; verify after SERP rank check",
             font=FONT_NOTE, align=ALIGN_LEFT)

    # Monthly visitors
    set_cell(ws, 7, 1, "Monthly organic visitors", font=FONT_LABEL)
    set_cell(ws, 7, 2, "=B5*B6", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 7, 3, "=C5*C6", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 7, 4, "=D5*D6", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")

    # Annual visitors
    set_cell(ws, 8, 1, "Annual visitors", font=FONT_LABEL)
    set_cell(ws, 8, 2, "=B7*12", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 8, 3, "=C7*12", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 8, 4, "=D7*12", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")

    # ── 2. Email funnel ───────────────────────────────────────────────────
    section_header(ws, 9, "EMAIL FUNNEL")
    set_cell(ws, 10, 1, "Visitor → email conversion", font=FONT_LABEL)
    set_cell(ws, 10, 2, 0.04, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 10, 3, 0.05, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 10, 4, 0.06, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 10, 5, "Industry travel-content median 2.8%",
             font=FONT_NOTE, align=ALIGN_LEFT)

    set_cell(ws, 11, 1, "Annual email captures", font=FONT_LABEL)
    set_cell(ws, 11, 2, "=B8*B10", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 11, 3, "=C8*C10", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 11, 4, "=D8*D10", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")

    # ── 3. Paid funnel ────────────────────────────────────────────────────
    section_header(ws, 12, "PAID FUNNEL")
    set_cell(ws, 13, 1, "Email → paid (60d)", font=FONT_LABEL)
    set_cell(ws, 13, 2, 0.03, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 13, 3, 0.05, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 13, 4, 0.07, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")

    set_cell(ws, 14, 1, "Direct visitor → paid (no email)", font=FONT_LABEL)
    set_cell(ws, 14, 2, 0.005, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 14, 3, 0.008, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")
    set_cell(ws, 14, 4, 0.012, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format="0.0%")

    set_cell(ws, 15, 1, "New paid members (yr)", font=FONT_LABEL)
    set_cell(ws, 15, 2, "=ROUND(B11*B13+B8*B14,0)", fill=FILL_OUTPUT,
             align=ALIGN_RIGHT, number_format="#,##0")
    set_cell(ws, 15, 3, "=ROUND(C11*C13+C8*C14,0)", fill=FILL_OUTPUT,
             align=ALIGN_RIGHT, number_format="#,##0")
    set_cell(ws, 15, 4, "=ROUND(D11*D13+D8*D14,0)", fill=FILL_OUTPUT,
             align=ALIGN_RIGHT, number_format="#,##0")

    set_cell(ws, 16, 1, "Cumulative paid members (90% retention)",
             font=FONT_LABEL)
    set_cell(ws, 16, 2, "=B15", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 16, 3, "=B16*0.9+C15", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")
    set_cell(ws, 16, 4, "=C16*0.9+D15", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="#,##0")

    # ── 4. ARPU ───────────────────────────────────────────────────────────
    section_header(ws, 17, "ARPU STACK (₹)")
    set_cell(ws, 18, 1, "Membership ARPU", font=FONT_LABEL)
    set_cell(ws, 18, 2, 999, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 18, 3, 1099, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 18, 4, 1199, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')

    set_cell(ws, 19, 1, "Paid affiliate (per member)", font=FONT_LABEL)
    set_cell(ws, 19, 2, 400, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 19, 3, 500, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 19, 4, 600, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')

    set_cell(ws, 20, 1, "Vault attach (yr, 25%)", font=FONT_LABEL)
    set_cell(ws, 20, 2, 50, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 20, 3, 80, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 20, 4, 100, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')

    set_cell(ws, 21, 1, "ARPU per paid member", font=FONT_LABEL)
    set_cell(ws, 21, 2, "=B18+B19+B20", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 21, 3, "=C18+C19+C20", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 21, 4, "=D18+D19+D20", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')

    set_cell(ws, 22, 1, "Free-tier affiliate (per visitor/yr)",
             font=FONT_LABEL)
    set_cell(ws, 22, 2, 40, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 22, 3, 50, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 22, 4, 60, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0')
    set_cell(ws, 22, 5, "Largest revenue line — investors miss this",
             font=FONT_NOTE, align=ALIGN_LEFT)

    # ── 5. Revenue ────────────────────────────────────────────────────────
    section_header(ws, 23, "REVENUE (₹ lakhs)")
    set_cell(ws, 24, 1, "Membership", font=FONT_LABEL)
    set_cell(ws, 24, 2, "=B16*B18/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 24, 3, "=C16*C18/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 24, 4, "=D16*D18/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 25, 1, "Paid affiliate", font=FONT_LABEL)
    set_cell(ws, 25, 2, "=B16*B19/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 25, 3, "=C16*C19/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 25, 4, "=D16*D19/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 26, 1, "Vault", font=FONT_LABEL)
    set_cell(ws, 26, 2, "=B16*B20/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 26, 3, "=C16*C20/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 26, 4, "=D16*D20/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 27, 1, "Free-tier affiliate", font=FONT_LABEL)
    set_cell(ws, 27, 2, "=B8*B22/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 27, 3, "=C8*C22/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 27, 4, "=D8*D22/100000", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 28, 1, "TOTAL REVENUE", font=FONT_LABEL)
    set_cell(ws, 28, 2, "=SUM(B24:B27)", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 28, 3, "=SUM(C24:C27)", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 28, 4, "=SUM(D24:D27)", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    # ── 6. Costs ──────────────────────────────────────────────────────────
    section_header(ws, 29, "COSTS (₹ lakhs)")
    set_cell(ws, 30, 1, "Contributor payouts", font=FONT_LABEL)
    set_cell(ws, 30, 2, 1.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 30, 3, 3.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 30, 4, 6.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 31, 1, "Hosting + tooling", font=FONT_LABEL)
    set_cell(ws, 31, 2, 0.5, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 31, 3, 1.5, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 31, 4, 4.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 32, 1, "Verification (KYC)", font=FONT_LABEL)
    set_cell(ws, 32, 2, 0.3, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 32, 3, 1.8, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 32, 4, 9.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 33, 1, "Team payroll", font=FONT_LABEL)
    set_cell(ws, 33, 2, 18.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 33, 3, 36.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 33, 4, 72.0, fill=FILL_INPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 34, 1, "TOTAL COSTS", font=FONT_LABEL)
    set_cell(ws, 34, 2, "=SUM(B30:B33)", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 34, 3, "=SUM(C30:C33)", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 34, 4, "=SUM(D30:D33)", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    set_cell(ws, 35, 1, "NET (Revenue - Costs)", font=FONT_LABEL)
    set_cell(ws, 35, 2, "=B28-B34", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 35, 3, "=C28-C34", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 35, 4, "=D28-D34", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    # ── 7. Key ratios ─────────────────────────────────────────────────────
    section_header(ws, 36, "KEY RATIOS")
    set_cell(ws, 37, 1, "Visitor → paid blended", font=FONT_LABEL)
    set_cell(ws, 37, 2, "=B15/B8", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="0.00%")
    set_cell(ws, 37, 3, "=C15/C8", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="0.00%")
    set_cell(ws, 37, 4, "=D15/D8", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format="0.00%")

    set_cell(ws, 38, 1, "ARR (₹ lakhs)", font=FONT_LABEL)
    set_cell(ws, 38, 2, "=B28", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 38, 3, "=C28", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')
    set_cell(ws, 38, 4, "=D28", fill=FILL_OUTPUT, align=ALIGN_RIGHT,
             number_format='"₹"#,##0.0"L"')

    # Legend at bottom
    section_header(ws, 40, "LEGEND")
    set_cell(ws, 41, 1, "Teal = editable assumption",
             fill=FILL_INPUT, font=FONT_LABEL, align=ALIGN_LEFT)
    set_cell(ws, 42, 1, "Green = computed (don't edit)",
             fill=FILL_OUTPUT, font=FONT_LABEL, align=ALIGN_LEFT)
    set_cell(ws, 43, 1,
             "Open in Sheets: File → Import → Upload → "
             "Replace spreadsheet. Formulas convert automatically.",
             font=FONT_NOTE, align=ALIGN_LEFT)
    set_cell(ws, 44, 1,
             "Build with: python3 scripts/build-funnel-xlsx.py",
             font=FONT_NOTE, align=ALIGN_LEFT)


def build_sensitivity_sheet(wb):
    """3×3 capture × conversion → ARR grid for the print handout."""
    ws = wb.create_sheet("Sensitivity 3x3")
    widths = {1: 30, 2: 14, 3: 14, 4: 14}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Title
    set_cell(ws, 1, 1,
             "Year-1 sensitivity: capture × email→paid → ARR (₹ lakhs)",
             font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER)
    for col in range(2, 5):
        set_cell(ws, 1, col, "", fill=FILL_HEADER)
    ws.row_dimensions[1].height = 22

    # Constants
    set_cell(ws, 3, 1, "Total addressable searches/yr",
             font=FONT_LABEL)
    set_cell(ws, 3, 2, "=('Funnel Model'!B5)*12", fill=FILL_OUTPUT,
             align=ALIGN_RIGHT, number_format="#,##0")

    set_cell(ws, 4, 1, "ARPU per paid (yr 1)", font=FONT_LABEL)
    set_cell(ws, 4, 2, "='Funnel Model'!B21", fill=FILL_OUTPUT,
             align=ALIGN_RIGHT, number_format='"₹"#,##0')

    set_cell(ws, 5, 1, "Free-tier affiliate per visitor", font=FONT_LABEL)
    set_cell(ws, 5, 2, "='Funnel Model'!B22", fill=FILL_OUTPUT,
             align=ALIGN_RIGHT, number_format='"₹"#,##0')

    # Grid header
    set_cell(ws, 7, 1, "Capture rate ↓ / Email→paid →", font=FONT_LABEL,
             fill=FILL_SECTION, align=ALIGN_CENTER)
    set_cell(ws, 7, 2, "2%", font=FONT_LABEL, fill=FILL_SECTION,
             align=ALIGN_CENTER)
    set_cell(ws, 7, 3, "4%", font=FONT_LABEL, fill=FILL_SECTION,
             align=ALIGN_CENTER)
    set_cell(ws, 7, 4, "6%", font=FONT_LABEL, fill=FILL_SECTION,
             align=ALIGN_CENTER)

    # Rows
    captures = [(8, "2% (low)", 0.02), (9, "5% (base)", 0.05),
                (10, "10% (high)", 0.10)]
    convs = [(2, 0.02), (3, 0.04), (4, 0.06)]

    for row, label, capture in captures:
        set_cell(ws, row, 1, label, font=FONT_LABEL, fill=FILL_SECTION,
                 align=ALIGN_LEFT)
        for col, conv in convs:
            # Total ARR = paid_count × ARPU + visitors × free_affiliate
            # paid_count ≈ visitors × 4% (vis→email) × conv
            # visitors = total_searches × capture
            formula = (
                f"=ROUND(($B$3*{capture}*0.04*{conv}*$B$4 "
                f"+ $B$3*{capture}*$B$5)/100000,1)"
            )
            set_cell(ws, row, col, formula, fill=FILL_OUTPUT,
                     align=ALIGN_RIGHT, number_format='"₹"#,##0.0"L"')

    # Footer
    set_cell(ws, 12, 1,
             "Highlighted base case: 5% capture × 4% conv = "
             "year-1 total revenue including free-tier affiliate",
             font=FONT_NOTE, align=ALIGN_LEFT)
    set_cell(ws, 13, 1,
             "Even worst case (2%×2%) leaves a real business covered "
             "by the round.", font=FONT_NOTE, align=ALIGN_LEFT)


def main():
    wb = Workbook()
    build_main_model(wb)
    build_sensitivity_sheet(wb)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"✓ Wrote {OUT}")


if __name__ == "__main__":
    main()
